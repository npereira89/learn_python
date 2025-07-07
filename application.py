import tkinter as tk
from tkinter import ttk, filedialog, messagebox, font
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from datetime import datetime
from dateutil.relativedelta import relativedelta

def load_file():
    try:
        file = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        return file
    except Exception as e:
        messagebox.showerror("ERROR", f"Error loading Excel file: {e}")


def clear_window():
    for widget in windows.winfo_children():
        if widget not in (menubar, menu, menu_2):
            widget.destroy()


def form_insert_data():
    frm_insert.pack(fill="both", expand=True)

def on_click_data(tree, file, sheet):
    if file:
        for widget in windows.winfo_children():
            if widget not in (menubar, menu, menu_2, tree, frm_insert):
                widget.destroy()
        selected_item = tree.selection()
        if selected_item:
            item_id = selected_item[0]
            value_cell = sheet.cell(row=int(item_id), column=2).value

            frm_update = tk.Frame(windows, padx=50, pady=25)
            frm_update.pack(expand=True, fill="both")

            label_value_upgrade = tk.Label(frm_update, text="Value (€):", fg="black", font=font.Font(weight="bold"))
            label_value_upgrade.pack()
            # Create the entry widget
            update_value = tk.Entry(frm_update)
            update_value.pack()

            update_button = tk.Button(frm_update, text="OK", width=4, height=2,
                                      command=lambda: update_data_excel(tree, value_cell, int(update_value.get()),
                                                                        item_id, file, frm_update),
                                      fg="blue", font=font.Font(weight="bold"))

            update_button.pack()


def update_data_excel(tree_updt, cell_value, value_upd, id_row, file_xlsx, frm_update):
    workbook = load_workbook(file_xlsx)
    sheet = workbook.active

    if value_upd == '' or value_upd == ' ' or value_upd == 0:
        messagebox.showwarning("WARNING", "The value updated be above 0")
    else:
        tree_updt.set(id_row, column=1, value=cell_value - value_upd)
        sheet.cell(row=int(id_row), column=2).value = int(cell_value - value_upd)

        if sheet.cell(row=int(id_row), column=2).value == 0:
            tree_updt.delete(id_row)
            sheet.delete_rows(int(id_row), amount=1)

        workbook.save(file_xlsx)
        workbook.close()
    frm_update.destroy()
    on_click_data(tree_updt, file_xlsx, sheet)

def load_excel_data():
    frm_insert.forget()
    file_xlsx = load_file()
    if file_xlsx:
        workbook = load_workbook(file_xlsx)
        sheet = workbook.active

        label = tk.Label(text="Tabela Certificados Aforro\n", font=("Courier New", 13, "bold underline"))
        label.pack()

        # Create a Treeview widget
        tree = ttk.Treeview(selectmode='browse')
        tree.pack(expand=True, fill="both")
        # # Get column names from the first row
        columns = list(sheet.iter_rows(values_only=True))[1]

        # Configure columns
        tree["columns"] = columns
        tree["show"] = "headings"
        # Set column headings
        for col in columns:
            tree.heading(col, text=col)

        # Insert data into the Treeview
        for row_num, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
            tree.insert("", "end", iid=row_num, values=row)

        messagebox.showinfo("SUCCESS", "All information was loaded.")
        tree.bind("<<TreeviewSelect>>", lambda event: on_click_data(tree, file_xlsx, sheet))


def save_excel_info(value_invest):
    file = load_file()
    if file:
        wb = load_workbook(file)
        sheet_ranges = wb.active

        now = datetime.now()
        new_date = now + relativedelta(months=3)

        if value_invest.isalpha() or value_invest == '':
            messagebox.showwarning("WARNING", "The invest must be above 0 and doesn't a string value")

        data = [
            [now.strftime("%d/%m/%Y"), int(value_invest), new_date.strftime("%d/%m/%Y"), 0.0275, "FALSE"]
        ]

        for row in range(2, len(data) + 2):
            max_row = sheet_ranges.max_row + 1
            for col in range(1, len(data[0]) + 1):
                sheet_ranges.cell(row=max_row, column=col).value = data[row - 2][col - 1]
                cell = sheet_ranges.cell(row=max_row, column=col)

                # Formatting borders data
                cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'),
                                     right=Side(style='thin'))

                # Formatting cells
                if col == 2 and cell.row == max_row:
                    cell.number_format = '#,##0€'
                elif col == 4:  # Format tax applied
                    cell.value = round(cell.value * 100, 2)

        wb.save(file)
        messagebox.showinfo("SUCCESS", "The invest was added with success!")
        wb.close()

# Create the main window
windows = tk.Tk()
windows.title("AforroApp")
windows.iconbitmap("python_icon.ico")
windows.geometry("800x600")

# Create a menu bar
menubar = tk.Menu()
windows.config(menu=menubar)

# Create entries menu
menu = tk.Menu(menubar, tearoff=0)
menu_2 = tk.Menu(menubar, tearoff=0)

# Menu open and load data Excel file and insert new data
menubar.add_cascade(label="File", menu=menu)
menu.add_command(label="Open", command=load_excel_data)
menu.add_command(label="Insert data", command=form_insert_data)
menu.add_command(label="Close Frames", command=clear_window)

# Menu close menu window
menubar.add_cascade(label="Windows", menu=menu_2)
menu_2.add_command(label="Exit", command=windows.quit)

# Frame to insert investment value
frm_insert = tk.Frame(windows, padx=50, pady=25)
frm_insert.pack(expand=True, fill="both")

tk.Label(frm_insert, text="Value (€):", fg="black", font=font.Font(weight="bold")).grid(row=0, column=0, padx=5,
                                                                                        pady=5)
value_entry = tk.Entry(frm_insert)
value_entry.grid(row=0, column=1, padx=5, pady=5)
tk.Button(frm_insert, text="OK", width=4, height=2, command=lambda: save_excel_info(value_entry.get()), fg="blue",
          font=font.Font(weight="bold")).grid(row=0, column=2, padx=5, pady=5)

frm_insert.forget()

windows.mainloop()